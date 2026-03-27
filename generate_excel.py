import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import json

# EPA Standard AQI Colors
AQI_COLORS = {
    "Good":                       ("00E400", "000000"),
    "Moderate":                   ("FFFF00", "000000"),
    "Unhealthy for Sensitive":    ("FF7E00", "000000"),
    "Unhealthy":                  ("FF0000", "FFFFFF"),
    "Very Unhealthy":             ("8F3F97", "FFFFFF"),
    "Hazardous":                  ("7E0023", "FFFFFF"),
}

def get_aqi_styles(aqi):
    """Return (Fill, FontColor) based on EPA standards."""
    if pd.isna(aqi):
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), "000000"
    
    val = int(aqi)
    if val <= 50:
        fill_hex, font_hex = AQI_COLORS["Good"]
    elif val <= 100:
        fill_hex, font_hex = AQI_COLORS["Moderate"]
    elif val <= 150:
        fill_hex, font_hex = AQI_COLORS["Unhealthy for Sensitive"]
    elif val <= 200:
        fill_hex, font_hex = AQI_COLORS["Unhealthy"]
    elif val <= 300:
        fill_hex, font_hex = AQI_COLORS["Very Unhealthy"]
    else:
        fill_hex, font_hex = AQI_COLORS["Hazardous"]
        
    return PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid"), font_hex

def generate_styled_excel():
    csv_path = "models/validation_predictions.csv"
    json_path = "models/validation_results.json"
    output_path = "AQI_Validation_Evidence.xlsx"
    
    if not os.path.exists(csv_path):
        print(f"Error: {csv_path} not found.")
        return

    print("Generating comprehensive professional Excel evidence...")
    
    # 1. Load Data
    df = pd.read_csv(csv_path)
    df = df.sort_values(['horizon_h', 'timestamp'])
    
    summary_data = []
    if os.path.exists(json_path):
        with open(json_path, 'r') as f:
            summary_data = json.load(f)

    # 2. Setup Workbook
    wb = openpyxl.Workbook()
    
    # --- SHEET 1: EXECUTIVE SUMMARY ---
    ws_sum = wb.active
    ws_sum.title = "Executive Summary"
    
    bold_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    title_cell = ws_sum.cell(row=1, column=1, value="AQI Model Performance Summary (STEM Fair Submission)")
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    
    headers = ["Horizon", "Sample Days", "Model MAE", "Baseline MAE", "Skill Score", "R² Score", "Coverage %"]
    for col, text in enumerate(headers, 1):
        cell = ws_sum.cell(row=3, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, data in enumerate(summary_data, 4):
        ws_sum.cell(row=row_idx, column=1, value=f"{data['horizon_h']} Hours").alignment = center_align
        ws_sum.cell(row=row_idx, column=2, value=data['n_folds']).alignment = center_align
        ws_sum.cell(row=row_idx, column=3, value=round(data['mean_mae'], 2)).alignment = center_align
        ws_sum.cell(row=row_idx, column=4, value=round(data['baseline_mae'], 2)).alignment = center_align
        
        # Skill Score highlighting
        ss_cell = ws_sum.cell(row=row_idx, column=5, value=f"{round(data['skill_score']*100, 1)}%")
        ss_cell.alignment = center_align
        if data['skill_score'] > 0:
            ss_cell.font = Font(color="008000", bold=True) # Green
        
        ws_sum.cell(row=row_idx, column=6, value=round(data['r2_score'], 3)).alignment = center_align
        
        cov_cell = ws_sum.cell(row=row_idx, column=7, value=f"{data['mean_coverage']}%")
        cov_cell.alignment = center_align
        if data['mean_coverage'] >= 90:
            cov_cell.font = Font(color="008000", bold=True)

    # --- SHEET 2: VALIDATION LOG ---
    ws_log = wb.create_sheet("Hourly Predictions Log")
    log_headers = ["Timestamp (PT)", "Actual AQI", "Predicted AQI", "Horizon", "Error (Abs)"]
    for col, text in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = center_align

    for row_idx, row_data in enumerate(df.values, 2):
        ts, act, prd, hor = row_data
        err = abs(int(act) - int(prd))
        vals = [ts, act, prd, f"{hor}h", err]
        
        for col_idx, val in enumerate(vals, 1):
            cell = ws_log.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in [2, 3]: # Style AQI cells
                fill, font_color = get_aqi_styles(val)
                cell.fill = fill
                cell.font = Font(color=font_color, bold=True)
                cell.alignment = center_align

    # --- SHEET 3: SCIENTIFIC METHODOLOGY ---
    ws_info = wb.create_sheet("About This Data")
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 80
    
    info = [
        ("Metric", "Definition"),
        ("Model MAE", "Mean Absolute Error: The average 'points off' the prediction was from the sensor."),
        ("Skill Score", "The % improvement of this model over a 'Persistence' baseline (predicting tomorrow is like today)."),
        ("R² Score", "The proportion of variance explained by the model (Correlation)."),
        ("Coverage %", "Percentage of time the actual value fell within the model's 95% Confidence Interval."),
        ("", ""),
        ("Science Note", "This model utilizes Gradient Boosting (LightGBM) with a Huber Loss function."),
        ("Huber Loss", "Optimizes for accuracy while minimizing the impact of outliers (like sudden smoke spikes)."),
        ("AOD Features", "Satellite Aerosol Optical Depth data is integrated to detect incoming smoke plumes."),
    ]
    
    for r_idx, (key, val) in enumerate(info, 1):
        k_cell = ws_info.cell(row=r_idx, column=1, value=key)
        v_cell = ws_info.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            k_cell.font = Font(bold=True)
            v_cell.font = Font(bold=True)
        k_cell.alignment = Alignment(vertical="top")
        v_cell.alignment = Alignment(wrap_text=True)

    # Global Column Auto-sizing
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column].width = max(max_length + 2, 10)

    # Save
    try:
        wb.save(output_path)
        print(f"Professional Excel evidence saved → {output_path}")
    except PermissionError:
        print(f"CRITICAL ERROR: Could not save {output_path}. Please close the file if it is open in Excel!")

if __name__ == "__main__":
    generate_styled_excel()
